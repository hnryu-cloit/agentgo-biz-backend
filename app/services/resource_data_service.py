from __future__ import annotations

import uuid
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable, Optional
import unicodedata

from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.models.resource_data import (
    BoPointUsageSnapshot,
    DodoPointSnapshot,
    MenuLineupSnapshot,
    PosDailySalesSnapshot,
    ReceiptTransactionSnapshot,
    ResourceStore,
)

SOURCE_SPECS = {
    "pos_daily_sales": {
        "label": "POS 일자별 매출",
        "description": "POS 원천에서 추출한 일자별 매출 집계 데이터",
        "directory_name": "다이닝본부_POS_RawData",
    },
    "bo_point_usage": {
        "label": "BO 포인트 사용",
        "description": "BO 기준 포인트/결제 수단 포함 일자별 매출 데이터",
        "directory_name": "다이닝본부_BO포인트사용_RawData",
    },
    "dodo_point": {
        "label": "도도포인트",
        "description": "도도포인트 기준 고객별 포인트 적립/사용 내역",
        "directory_name": "다이닝본부_도도포인트_RawData",
    },
    "receipt_listing": {
        "label": "영수증 목록",
        "description": "영수증 단위 거래 로그 데이터",
        "directory_name": "다이닝본부_영수증목록리스트_RawData",
    },
    "menu_lineup": {
        "label": "메뉴 라인업",
        "description": "매장별 메뉴 구성 및 원가표 데이터",
        "directory_name": "전점메뉴라인업",
    },
}

# 메뉴 라인업 시트명 → 정규 store_key 매핑
# (시트명이 브랜드명과 다른 경우 여기에 추가)
MENU_SHEET_STORE_KEY_MAP: dict[str, str] = {
    "천호점": "크리스탈제이드",
}


class ResourceDataService:
    def __init__(self, resource_dir: Optional[str] = None):
        self.base_dir = Path(resource_dir or settings.RESOURCE_DATA_DIR)

    def list_catalog(self) -> dict[str, Any]:
        sources = []
        for source_kind, spec in SOURCE_SPECS.items():
            stores = self._catalog_for_source(source_kind)
            sources.append(
                {
                    "source_kind": source_kind,
                    "label": spec["label"],
                    "description": spec["description"],
                    "stores": stores,
                }
            )
        return {"sources": sources}

    def get_dataset(self, source_kind: str, store_key: str, limit: int = 50) -> dict[str, Any]:
        self._validate_source_kind(source_kind)
        store_key = self._normalize_key(store_key)
        if source_kind == "menu_lineup":
            return self._get_menu_dataset(store_key, limit)
        return self._get_tabular_dataset(source_kind, store_key, limit)

    async def list_catalog_from_db(self, db: AsyncSession) -> dict[str, Any]:
        sources = []
        for source_kind, spec in SOURCE_SPECS.items():
            stores = await self._catalog_for_source_from_db(db, source_kind)
            sources.append(
                {
                    "source_kind": source_kind,
                    "label": spec["label"],
                    "description": spec["description"],
                    "stores": stores,
                }
            )
        return {"sources": sources}

    async def get_dataset_from_db(
        self,
        db: AsyncSession,
        source_kind: str,
        store_key: str,
        limit: int = 50,
    ) -> dict[str, Any]:
        self._validate_source_kind(source_kind)
        normalized_store_key = self._normalize_key(store_key)
        if source_kind == "pos_daily_sales":
            return await self._get_pos_dataset_from_db(db, normalized_store_key, limit)
        if source_kind == "bo_point_usage":
            return await self._get_bo_dataset_from_db(db, normalized_store_key, limit)
        if source_kind == "dodo_point":
            return await self._get_dodo_dataset_from_db(db, normalized_store_key, limit)
        if source_kind == "receipt_listing":
            return await self._get_receipt_dataset_from_db(db, normalized_store_key, limit)
        return await self._get_menu_dataset_from_db(db, normalized_store_key, limit)

    async def import_dataset(
        self,
        db: AsyncSession,
        source_kind: str,
        store_key: Optional[str] = None,
    ) -> int:
        self._validate_source_kind(source_kind)
        normalized_store_key = self._normalize_key(store_key) if store_key else None
        if source_kind == "menu_lineup":
            return await self._import_menu_dataset(db, normalized_store_key)
        return await self._import_tabular_dataset(db, source_kind, normalized_store_key)

    async def _catalog_for_source_from_db(self, db: AsyncSession, source_kind: str) -> list[dict[str, Any]]:
        result = await db.execute(
            select(ResourceStore)
            .where(ResourceStore.source_kind == source_kind)
            .order_by(ResourceStore.store_name.asc())
        )
        stores = list(result.scalars().all())
        response = []
        for store in stores:
            metadata = store.metadata_json or {}
            date_start = metadata.get("date_start")
            date_end = metadata.get("date_end")
            response.append(
                {
                    "store_key": store.store_key,
                    "latest_file_name": store.latest_file_name,
                    "file_count": int(metadata.get("file_count", 1)),
                    "date_start": date_start,
                    "date_end": date_end,
                }
            )
        return response

    async def _get_pos_dataset_from_db(self, db: AsyncSession, store_key: str, limit: int) -> dict[str, Any]:
        rows = list(
            (
                await db.execute(
                    select(PosDailySalesSnapshot)
                    .where(PosDailySalesSnapshot.store_key == store_key)
                    .order_by(PosDailySalesSnapshot.sales_date.desc())
                    .limit(limit)
                )
            ).scalars().all()
        )
        if not rows:
            raise ValueError(f"No imported dataset found for source_kind=pos_daily_sales, store_key={store_key}")
        headers = [
            "sales_date", "store_name", "quantity", "guest_count", "guest_avg_spend",
            "receipt_count", "receipt_avg_spend", "gross_sales_amount", "refund_amount",
            "total_sales_amount", "discount_amount", "net_sales_amount",
        ]
        return {
            "source_kind": "pos_daily_sales",
            "store_key": store_key,
            "headers": headers,
            "rows": [
                {
                    "sales_date": row.sales_date.isoformat(),
                    "store_name": row.store_name,
                    "quantity": row.quantity,
                    "guest_count": row.guest_count,
                    "guest_avg_spend": row.guest_avg_spend,
                    "receipt_count": row.receipt_count,
                    "receipt_avg_spend": row.receipt_avg_spend,
                    "gross_sales_amount": row.gross_sales_amount,
                    "refund_amount": row.refund_amount,
                    "total_sales_amount": row.total_sales_amount,
                    "discount_amount": row.discount_amount,
                    "net_sales_amount": row.net_sales_amount,
                }
                for row in rows
            ],
            "summary": await self._build_db_summary(
                db,
                PosDailySalesSnapshot,
                store_key,
                "sales_date",
                total_field="total_sales_amount",
            ),
        }

    async def _get_bo_dataset_from_db(self, db: AsyncSession, store_key: str, limit: int) -> dict[str, Any]:
        rows = list(
            (
                await db.execute(
                    select(BoPointUsageSnapshot)
                    .where(BoPointUsageSnapshot.store_key == store_key)
                    .order_by(BoPointUsageSnapshot.sales_date.desc())
                    .limit(limit)
                )
            ).scalars().all()
        )
        if not rows:
            raise ValueError(f"No imported dataset found for source_kind=bo_point_usage, store_key={store_key}")
        headers = [
            "sales_date", "store_name", "weekday_label", "receipt_count", "customer_count",
            "gross_sales_amount", "sales_amount", "payment_total_amount", "discount_amount", "refund_amount",
        ]
        return {
            "source_kind": "bo_point_usage",
            "store_key": store_key,
            "headers": headers,
            "rows": [
                {
                    "sales_date": row.sales_date.isoformat(),
                    "store_name": row.store_name,
                    "weekday_label": row.weekday_label,
                    "receipt_count": row.receipt_count,
                    "customer_count": row.customer_count,
                    "gross_sales_amount": row.gross_sales_amount,
                    "sales_amount": row.sales_amount,
                    "payment_total_amount": row.payment_total_amount,
                    "discount_amount": row.discount_amount,
                    "refund_amount": row.refund_amount,
                }
                for row in rows
            ],
            "summary": await self._build_db_summary(
                db,
                BoPointUsageSnapshot,
                store_key,
                "sales_date",
                total_field="payment_total_amount",
            ),
        }

    async def _get_receipt_dataset_from_db(self, db: AsyncSession, store_key: str, limit: int) -> dict[str, Any]:
        rows = list(
            (
                await db.execute(
                    select(ReceiptTransactionSnapshot)
                    .where(ReceiptTransactionSnapshot.store_key == store_key)
                    .order_by(ReceiptTransactionSnapshot.sales_date.desc(), ReceiptTransactionSnapshot.sales_time.desc())
                    .limit(limit)
                )
            ).scalars().all()
        )
        if not rows:
            raise ValueError(f"No imported dataset found for source_kind=receipt_listing, store_key={store_key}")
        headers = [
            "sales_date", "sales_time", "store_name", "pos_name", "transaction_number",
            "sales_category", "transaction_type", "gross_amount", "discount_amount", "payment_amount",
        ]
        return {
            "source_kind": "receipt_listing",
            "store_key": store_key,
            "headers": headers,
            "rows": [
                {
                    "sales_date": row.sales_date.isoformat(),
                    "sales_time": row.sales_time.isoformat() if row.sales_time else None,
                    "store_name": row.store_name,
                    "pos_name": row.pos_name,
                    "transaction_number": row.transaction_number,
                    "sales_category": row.sales_category,
                    "transaction_type": row.transaction_type,
                    "gross_amount": row.gross_amount,
                    "discount_amount": row.discount_amount,
                    "payment_amount": row.payment_amount,
                }
                for row in rows
            ],
            "summary": await self._build_db_summary(
                db,
                ReceiptTransactionSnapshot,
                store_key,
                "sales_date",
                total_field="payment_amount",
            ),
        }

    async def _get_dodo_dataset_from_db(self, db: AsyncSession, store_key: str, limit: int) -> dict[str, Any]:
        rows = list(
            (
                await db.execute(
                    select(DodoPointSnapshot)
                    .where(DodoPointSnapshot.store_key == store_key)
                    .order_by(DodoPointSnapshot.event_at.desc())
                    .limit(limit)
                )
            ).scalars().all()
        )
        if not rows:
            raise ValueError(f"No imported dataset found for source_kind=dodo_point, store_key={store_key}")
        headers = ["event_at", "event_date", "store_name", "customer_masked", "point_type", "point_amount"]
        return {
            "source_kind": "dodo_point",
            "store_key": store_key,
            "headers": headers,
            "rows": [
                {
                    "event_at": row.event_at.isoformat() if row.event_at else None,
                    "event_date": row.event_date.isoformat(),
                    "store_name": row.store_name,
                    "customer_masked": row.customer_masked,
                    "point_type": row.point_type,
                    "point_amount": row.point_amount,
                }
                for row in rows
            ],
            "summary": await self._build_db_summary(
                db,
                DodoPointSnapshot,
                store_key,
                "event_date",
                total_field="point_amount",
            ),
        }

    async def _get_menu_dataset_from_db(self, db: AsyncSession, store_key: str, limit: int) -> dict[str, Any]:
        rows = list(
            (
                await db.execute(
                    select(MenuLineupSnapshot)
                    .where(MenuLineupSnapshot.store_key == store_key)
                    .order_by(MenuLineupSnapshot.row_number.asc())
                    .limit(limit)
                )
            ).scalars().all()
        )
        if not rows:
            raise ValueError(f"No imported dataset found for source_kind=menu_lineup, store_key={store_key}")
        headers = ["row_number", "menu_category", "menu_name", "sales_price", "cost_amount", "cost_rate"]
        return {
            "source_kind": "menu_lineup",
            "store_key": store_key,
            "headers": headers,
            "rows": [
                {
                    "row_number": row.row_number,
                    "menu_category": row.menu_category,
                    "menu_name": row.menu_name,
                    "sales_price": row.sales_price,
                    "cost_amount": row.cost_amount,
                    "cost_rate": row.cost_rate,
                }
                for row in rows
            ],
            "summary": await self._build_db_summary(
                db,
                MenuLineupSnapshot,
                store_key,
                None,
                total_field=None,
            ),
        }

    def _catalog_for_source(self, source_kind: str) -> list[dict[str, Any]]:
        if source_kind == "menu_lineup":
            return self._catalog_for_menu_lineup()

        store_files = defaultdict(list)
        for file_path in self._iter_source_files(source_kind):
            store_files[self._resolve_store_key(source_kind, file_path)].append(file_path)

        stores = []
        for store_key, files in sorted(store_files.items()):
            latest_file = max(files, key=lambda item: item.name)
            period_range = self._extract_date_range_from_filename(latest_file.name)
            stores.append(
                {
                    "store_key": store_key,
                    "latest_file_name": latest_file.name,
                    "file_count": len(files),
                    "date_start": period_range[0],
                    "date_end": period_range[1],
                }
            )
        return stores

    def _catalog_for_menu_lineup(self) -> list[dict[str, Any]]:
        workbook_path = self._get_menu_workbook_path()
        sheet_names = self._load_sheet_names(workbook_path)
        return [
            {
                "store_key": MENU_SHEET_STORE_KEY_MAP.get(
                    self._normalize_key(sheet_name),
                    self._normalize_key(sheet_name),
                ),
                "latest_file_name": workbook_path.name,
                "file_count": 1,
                "date_start": None,
                "date_end": None,
            }
            for sheet_name in sheet_names
        ]

    def _get_tabular_dataset(self, source_kind: str, store_key: str, limit: int) -> dict[str, Any]:
        files = sorted(
            self._iter_source_files(source_kind, store_key),
            key=lambda item: item.name,
            reverse=True,
        )
        if not files:
            raise ValueError(f"No files found for source_kind={source_kind}, store_key={store_key}")

        target_file = files[0]
        header, rows = self._load_first_sheet_rows(target_file)
        normalized_rows = [self._row_to_dict(header, row) for row in rows if self._is_data_row(row)]
        summary = self._build_tabular_summary(source_kind, normalized_rows, target_file.name)
        return {
            "source_kind": source_kind,
            "store_key": store_key,
            "headers": header,
            "rows": normalized_rows[:limit],
            "summary": summary,
        }

    def _get_menu_dataset(self, store_key: str, limit: int) -> dict[str, Any]:
        workbook_path = self._get_menu_workbook_path()
        workbook = self._load_workbook(workbook_path, read_only=True, data_only=True)
        normalized_sheet_map = {self._normalize_key(name): name for name in workbook.sheetnames}
        if store_key not in normalized_sheet_map:
            raise ValueError(f"Sheet not found for store_key={store_key}")

        sheet = workbook[normalized_sheet_map[store_key]]
        rows = []
        headers = ["row_number", "values"]
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [self._clean_cell(value) for value in row]
            if not any(values):
                continue
            rows.append({"row_number": idx, "values": values})
            if len(rows) >= limit:
                break

        return {
            "source_kind": "menu_lineup",
            "store_key": store_key,
            "headers": headers,
            "rows": rows,
            "summary": {
                "file_name": workbook_path.name,
                "sheet_name": normalized_sheet_map[store_key],
                "preview_row_count": len(rows),
            },
        }

    async def _import_tabular_dataset(
        self,
        db: AsyncSession,
        source_kind: str,
        store_key: Optional[str],
    ) -> int:
        imported_count = 0
        files = list(self._iter_source_files(source_kind, store_key))
        for file_path in files:
            header, rows = self._load_first_sheet_rows(file_path)
            records = [self._row_to_dict(header, row) for row in rows if self._is_data_row(row)]
            if not records:
                continue

            # POS/BO: 레코드 내 매장명 기준으로 매장별 그룹화
            # 도도포인트: 디렉토리명(폴더)을 store_key로 사용 (매장 이름이 풀네임이므로)
            # 영수증: 레코드 내 '매장' 컬럼 기준
            if source_kind in ("pos_daily_sales", "bo_point_usage"):
                store_name_field = "매장명"
            elif source_kind == "receipt_listing":
                store_name_field = "매장"
            else:
                store_name_field = None

            file_store_key = self._resolve_store_key(source_kind, file_path)
            by_store: dict[str, list] = {}
            for rec in records:
                if store_name_field:
                    raw_name = self._string_or_none(rec.get(store_name_field))
                    sk = self._normalize_key(raw_name) if raw_name else file_store_key
                else:
                    sk = file_store_key
                by_store.setdefault(sk, []).append(rec)

            for sk, recs in by_store.items():
                first = recs[0]
                store_name = self._string_or_none(first.get(store_name_field)) or sk
                date_field_key = "날짜" if source_kind == "dodo_point" else "매출일자"
                store_name = self._string_or_none(first.get(store_name_field)) if store_name_field else sk
                await self._upsert_resource_store(
                    db=db,
                    source_kind=source_kind,
                    store_key=sk,
                    external_store_code=self._string_or_none(first.get("매장코드")),
                    store_name=store_name or sk,
                    latest_file_name=file_path.name,
                    metadata_json={
                        "header": header,
                        "row_count": len(recs),
                        "file_count": 1,
                        "date_start": self._isoformat_date(
                            min(
                                (self._parse_date(record.get(date_field_key)) for record in recs if self._parse_date(record.get(date_field_key))),
                                default=None,
                            )
                        ),
                        "date_end": self._isoformat_date(
                            max(
                                (self._parse_date(record.get(date_field_key)) for record in recs if self._parse_date(record.get(date_field_key))),
                                default=None,
                            )
                        ),
                    },
                )

                if source_kind == "pos_daily_sales":
                    await db.execute(
                        delete(PosDailySalesSnapshot).where(
                            PosDailySalesSnapshot.source_file_name == file_path.name,
                            PosDailySalesSnapshot.store_key == sk,
                        )
                    )
                    for record in recs:
                        sales_date = self._parse_date(record.get("매출일자"))
                        if not sales_date:
                            continue
                        db.add(
                            PosDailySalesSnapshot(
                                id=str(uuid.uuid4()),
                                store_key=sk,
                                store_code=self._string_or_none(record.get("매장코드")),
                                store_name=self._string_or_none(record.get("매장명")) or sk,
                                sales_date=sales_date,
                                quantity=self._to_float(record.get("수량")),
                                guest_count=self._to_float(record.get("객수")),
                                guest_avg_spend=self._to_float(record.get("객단가")),
                                receipt_count=self._to_float(record.get("영수건수")),
                                receipt_avg_spend=self._to_float(record.get("영수단가")),
                                gross_sales_amount=self._to_float(record.get("총판매금액")),
                                refund_amount=self._to_float(record.get("총반품금액")),
                                total_sales_amount=self._to_float(record.get("총 매출 금액")),
                                discount_amount=self._to_float(record.get("할인 금액")),
                                net_sales_amount=self._to_float(record.get("순매출")),
                                sales_amount=self._to_float(record.get("매출 금액")),
                                cash_sales_amount=self._to_float(record.get("현금매출")),
                                card_sales_amount=self._to_float(record.get("카드매출")),
                                simple_payment_sales_amount=self._to_float(record.get("간편결제매출")),
                                giftcard_sales_amount=self._to_float(record.get("상품권매출")),
                                point_sales_amount=self._to_float(record.get("포인트매출")),
                                order_channel_sales_amount=self._to_float(record.get("오더주문매출")),
                                source_file_name=file_path.name,
                            )
                        )
                        imported_count += 1
                elif source_kind == "bo_point_usage":
                    await db.execute(
                        delete(BoPointUsageSnapshot).where(
                            BoPointUsageSnapshot.source_file_name == file_path.name,
                            BoPointUsageSnapshot.store_key == sk,
                        )
                    )
                    for record in recs:
                        sales_date = self._parse_date(record.get("매출일자"))
                        if not sales_date:
                            continue
                        db.add(
                            BoPointUsageSnapshot(
                                id=str(uuid.uuid4()),
                                store_key=sk,
                                store_code=self._string_or_none(record.get("매장코드")),
                                store_name=self._string_or_none(record.get("매장명")) or sk,
                                sales_date=sales_date,
                                weekday_label=self._string_or_none(record.get("요일")),
                                receipt_count=self._to_float(record.get("영수건수")),
                                team_count=self._to_float(record.get("조수")),
                                team_avg_spend=self._to_float(record.get("조단가")),
                                customer_count=self._to_float(record.get("고객수")),
                                gross_sales_amount=self._to_float(record.get("총매출")),
                                sales_amount=self._to_float(record.get("매출")),
                                payment_total_amount=self._to_float(record.get("결제수단 총합")),
                                net_sales_vat_excluded=self._to_float(record.get("순매출액(VAT제외)")),
                                discount_amount=self._to_float(record.get("할인")),
                                service_discount_amount=self._to_float(record.get("서비스 할인")),
                                refund_amount=self._to_float(record.get("반품")),
                                other_sales_amount=self._to_float(record.get("타상매출")),
                                cash_amount=self._to_float(record.get("현금")),
                                credit_card_total_amount=self._to_float(record.get("신용카드(전체)")),
                                credit_card_pos_amount=self._to_float(record.get("신용카드(POS)")),
                                credit_card_external_amount=self._to_float(record.get("신용카드(POS 외)")),
                                source_file_name=file_path.name,
                            )
                        )
                        imported_count += 1
                elif source_kind == "receipt_listing":
                    await db.execute(
                        delete(ReceiptTransactionSnapshot).where(
                            ReceiptTransactionSnapshot.source_file_name == file_path.name,
                            ReceiptTransactionSnapshot.store_key == sk,
                        )
                    )
                    for record in recs:
                        sales_date = self._parse_date(record.get("매출일자"))
                        if not sales_date:
                            continue
                        db.add(
                            ReceiptTransactionSnapshot(
                                id=str(uuid.uuid4()),
                                store_key=sk,
                                store_name=self._string_or_none(record.get("매장")) or sk,
                                sales_date=sales_date,
                                sales_time=self._parse_time(record.get("시간")),
                                pos_name=self._string_or_none(record.get("POS")),
                                transaction_number=self._string_or_none(record.get("거래번호")),
                                sales_category=self._string_or_none(record.get("매출구분")),
                                transaction_type=self._string_or_none(record.get("거래종류")),
                                cashier_code=self._string_or_none(record.get("계산원")),
                                gross_amount=self._to_float(record.get("총매출금액")),
                                total_quantity=self._to_float(record.get("총수량")),
                                discount_amount=self._to_float(record.get("할인금액")),
                                payment_amount=self._to_float(record.get("결제금액")),
                                order_number=self._string_or_none(record.get("주문번호")),
                                section_code=self._string_or_none(record.get("섹션코드")),
                                table_name=self._string_or_none(record.get("테이블명")),
                                table_staff=self._string_or_none(record.get("테이블담당자")),
                                e_receipt_issued=self._string_or_none(record.get("전자영수증\n발행여부")),
                                source_file_name=file_path.name,
                            )
                        )
                        imported_count += 1
                elif source_kind == "dodo_point":
                    await db.execute(
                        delete(DodoPointSnapshot).where(
                            DodoPointSnapshot.source_file_name == file_path.name,
                            DodoPointSnapshot.store_key == sk,
                        )
                    )
                    for record in recs:
                        raw_dt = record.get("날짜")
                        event_dt = raw_dt if isinstance(raw_dt, datetime) else None
                        event_date = self._parse_date(raw_dt)
                        if not event_date:
                            continue
                        db.add(
                            DodoPointSnapshot(
                                id=str(uuid.uuid4()),
                                store_key=sk,
                                store_name=self._string_or_none(record.get("매장 이름")) or sk,
                                event_at=event_dt,
                                event_date=event_date,
                                customer_masked=self._string_or_none(record.get("고객")),
                                customer_uuid=self._string_or_none(record.get("고객 UUID")),
                                point_type=self._string_or_none(record.get("종류")) or "",
                                point_amount=self._to_float(record.get("포인트")),
                                source_file_name=file_path.name,
                            )
                        )
                        imported_count += 1

        await db.commit()
        return imported_count

    async def _import_menu_dataset(self, db: AsyncSession, store_key: Optional[str]) -> int:
        workbook_path = self._get_menu_workbook_path()
        workbook = self._load_workbook(workbook_path, read_only=True, data_only=True)
        # store_key 기준으로 시트 찾기: 역방향 매핑 (크리스탈제이드 → 천호점)
        reverse_map = {v: k for k, v in MENU_SHEET_STORE_KEY_MAP.items()}
        normalized_sheet_map = {self._normalize_key(name): name for name in workbook.sheetnames}

        if store_key:
            # 매핑된 시트명 우선 조회, 없으면 store_key를 시트명으로 직접 사용
            sheet_name_key = reverse_map.get(store_key, store_key)
            actual_sheet = normalized_sheet_map.get(sheet_name_key)
            target_sheets = [actual_sheet] if actual_sheet else []
        else:
            target_sheets = list(workbook.sheetnames)

        imported_count = 0

        for sheet_name in target_sheets:
            if sheet_name not in workbook.sheetnames:
                continue
            normalized_sheet = self._normalize_key(sheet_name)
            canonical_store_key = MENU_SHEET_STORE_KEY_MAP.get(normalized_sheet, normalized_sheet)

            await db.execute(delete(MenuLineupSnapshot).where(MenuLineupSnapshot.sheet_name == sheet_name))
            await self._upsert_resource_store(
                db=db,
                source_kind="menu_lineup",
                store_key=canonical_store_key,
                external_store_code=None,
                store_name=sheet_name,
                latest_file_name=workbook_path.name,
                metadata_json={"sheet_name": sheet_name, "file_count": 1},
            )

            sheet = workbook[sheet_name]
            for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [self._clean_cell(value) for value in row]
                if not any(values):
                    continue
                db.add(
                    MenuLineupSnapshot(
                        id=str(uuid.uuid4()),
                        store_key=canonical_store_key,
                        sheet_name=sheet_name,
                        row_number=idx,
                        menu_category=self._string_or_none(values[1]) if len(values) > 1 else None,
                        menu_name=self._string_or_none(values[3]) if len(values) > 3 else None,
                        sales_price=self._to_float(values[4]) if len(values) > 4 else None,
                        cost_amount=self._to_float(values[5]) if len(values) > 5 else None,
                        cost_rate=self._to_float(values[6]) if len(values) > 6 else None,
                        row_payload={"values": values},
                        source_file_name=workbook_path.name,
                    )
                )
                imported_count += 1

        await db.commit()
        return imported_count

    async def _upsert_resource_store(
        self,
        db: AsyncSession,
        source_kind: str,
        store_key: str,
        external_store_code: Optional[str],
        store_name: str,
        latest_file_name: Optional[str],
        metadata_json: Optional[dict],
    ) -> None:
        await db.execute(
            delete(ResourceStore).where(
                ResourceStore.source_kind == source_kind,
                ResourceStore.store_key == store_key,
            )
        )
        db.add(
            ResourceStore(
                id=str(uuid.uuid4()),
                source_kind=source_kind,
                store_key=store_key,
                external_store_code=external_store_code,
                store_name=store_name,
                latest_file_name=latest_file_name,
                metadata_json=metadata_json,
            )
        )

    async def _build_db_summary(
        self,
        db: AsyncSession,
        model: Any,
        store_key: str,
        date_field_name: Optional[str],
        total_field: Optional[str],
    ) -> dict[str, Any]:
        summary: dict[str, Any] = {"row_count": 0}
        count_query = select(func.count()).select_from(model).where(model.store_key == store_key)
        summary["row_count"] = int((await db.execute(count_query)).scalar_one() or 0)

        if date_field_name:
            date_field = getattr(model, date_field_name)
            summary["date_start"] = (await db.execute(select(func.min(date_field)).where(model.store_key == store_key))).scalar_one_or_none()
            summary["date_end"] = (await db.execute(select(func.max(date_field)).where(model.store_key == store_key))).scalar_one_or_none()

        if total_field:
            metric_field = getattr(model, total_field)
            summary[f"{total_field}_total"] = round(
                float((await db.execute(select(func.coalesce(func.sum(metric_field), 0)).where(model.store_key == store_key))).scalar_one() or 0.0),
                2,
            )

        if "date_start" in summary and summary["date_start"] is not None:
            summary["date_start"] = summary["date_start"].isoformat()
        if "date_end" in summary and summary["date_end"] is not None:
            summary["date_end"] = summary["date_end"].isoformat()
        return summary

    def _load_first_sheet_rows(self, file_path: Path) -> tuple[list[str], list[list[Any]]]:
        workbook = self._load_workbook(file_path, read_only=False, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        row_iter = sheet.iter_rows(values_only=True)
        header = [self._string_or_none(value) or "" for value in next(row_iter)]
        rows = [list(row) for row in row_iter]
        return header, rows

    def _load_sheet_names(self, file_path: Path) -> list[str]:
        workbook = self._load_workbook(file_path, read_only=False, data_only=True)
        return list(workbook.sheetnames)

    def _load_workbook(self, file_path: Path, *, read_only: bool, data_only: bool):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "openpyxl is required to read Excel resource datasets. Install backend requirements before using data import features."
            ) from exc
        return load_workbook(file_path, read_only=read_only, data_only=data_only)

    def _build_tabular_summary(
        self,
        source_kind: str,
        rows: list[dict[str, Any]],
        file_name: str,
    ) -> dict[str, Any]:
        date_field = "날짜" if source_kind == "dodo_point" else "매출일자"
        date_values = [self._parse_date(row.get(date_field)) for row in rows]
        date_values = [value for value in date_values if value]
        summary: dict[str, Any] = {
            "file_name": file_name,
            "row_count": len(rows),
            "date_start": min(date_values) if date_values else None,
            "date_end": max(date_values) if date_values else None,
        }
        if source_kind == "pos_daily_sales":
            summary["net_sales_total"] = round(sum(self._to_float(row.get("순매출")) or 0 for row in rows), 2)
        elif source_kind == "bo_point_usage":
            summary["payment_total_amount"] = round(sum(self._to_float(row.get("결제수단 총합")) or 0 for row in rows), 2)
        elif source_kind == "receipt_listing":
            summary["payment_amount_total"] = round(sum(self._to_float(row.get("결제금액")) or 0 for row in rows), 2)
        elif source_kind == "dodo_point":
            summary["total_events"] = len(rows)
            summary["earn_count"] = sum(1 for row in rows if self._string_or_none(row.get("종류")) == "적립")
            summary["use_count"] = sum(1 for row in rows if self._string_or_none(row.get("종류")) == "사용")
            summary["unique_customers"] = len({row.get("고객 UUID") for row in rows if row.get("고객 UUID")})
        return summary

    def _row_to_dict(self, header: list[str], row: Iterable[Any]) -> dict[str, Any]:
        values = list(row)
        row_dict: dict[str, Any] = {}
        for idx, column in enumerate(header):
            if not column:
                continue
            row_dict[column] = self._clean_cell(values[idx] if idx < len(values) else None)
        return row_dict

    def _is_data_row(self, row: Iterable[Any]) -> bool:
        values = [self._clean_cell(value) for value in row]
        if not any(values):
            return False
        first_value = self._string_or_none(values[0])
        if first_value == "합계":
            return False
        return True

    def _iter_source_files(self, source_kind: str, store_key: Optional[str] = None) -> Iterable[Path]:
        source_dir = self._get_source_dir(SOURCE_SPECS[source_kind]["directory_name"])
        if source_kind == "menu_lineup":
            return [self._get_menu_workbook_path()]

        if store_key and source_kind != "receipt_listing":
            source_dir = self._find_child_dir(source_dir, store_key)
        if not source_dir.exists():
            return []
        files = sorted(source_dir.rglob("*.xlsx"))
        if source_kind == "receipt_listing" and store_key:
            return [
                file_path
                for file_path in files
                if self._normalize_key(self._resolve_store_key(source_kind, file_path)) == store_key
            ]
        return files

    def _get_menu_workbook_path(self) -> Path:
        source_dir = self._get_source_dir(SOURCE_SPECS["menu_lineup"]["directory_name"])
        files = sorted(source_dir.glob("*.xlsx"))
        if not files:
            raise ValueError("Menu lineup workbook not found")
        return files[0]

    def _get_source_dir(self, directory_name: str) -> Path:
        direct = self.base_dir / directory_name
        if direct.exists():
            return direct

        normalized_target = self._normalize_key(directory_name)
        for child in self.base_dir.iterdir():
            if child.is_dir() and self._normalize_key(child.name) == normalized_target:
                return child
        return direct

    def _find_child_dir(self, parent_dir: Path, child_name: str) -> Path:
        direct = parent_dir / child_name
        if direct.exists():
            return direct

        normalized_target = self._normalize_key(child_name)
        for child in parent_dir.iterdir():
            if child.is_dir() and self._normalize_key(child.name) == normalized_target:
                return child
        return direct

    def _extract_date_range_from_filename(self, file_name: str) -> tuple[Optional[date], Optional[date]]:
        date_candidates = []
        for part in file_name.replace(".xlsx", "").split("_"):
            parsed = self._parse_date(part)
            if parsed:
                date_candidates.append(parsed)
        if not date_candidates:
            return None, None
        if len(date_candidates) == 1:
            return date_candidates[0], date_candidates[0]
        return date_candidates[-2], date_candidates[-1]

    def _validate_source_kind(self, source_kind: str) -> None:
        if source_kind not in SOURCE_SPECS:
            raise ValueError(f"Unsupported source_kind={source_kind}")

    def _resolve_store_key(self, source_kind: str, file_path: Path) -> str:
        if source_kind == "receipt_listing":
            return self._normalize_key(file_path.stem.split("_")[0])
        return self._normalize_key(file_path.parent.name)

    def _normalize_key(self, value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        return unicodedata.normalize("NFC", value.strip())

    def _clean_cell(self, value: Any) -> Any:
        if isinstance(value, str):
            return value.strip()
        return value

    def _string_or_none(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        string_value = str(value).strip()
        return string_value or None

    def _to_float(self, value: Any) -> Optional[float]:
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            normalized = value.replace(",", "").strip()
            if not normalized:
                return None
            try:
                return float(normalized)
            except ValueError:
                return None
        return None

    def _parse_date(self, value: Any) -> Optional[date]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            normalized = value.strip()
            for pattern in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
                try:
                    return datetime.strptime(normalized, pattern).date()
                except ValueError:
                    continue
        return None

    def _parse_time(self, value: Any) -> Optional[time]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.time()
        if isinstance(value, time):
            return value
        if isinstance(value, str):
            normalized = value.strip()
            for pattern in ("%H:%M:%S", "%H:%M"):
                try:
                    return datetime.strptime(normalized, pattern).time()
                except ValueError:
                    continue
        return None

    def _isoformat_date(self, value: Optional[date]) -> Optional[str]:
        if value is None:
            return None
        return value.isoformat()
